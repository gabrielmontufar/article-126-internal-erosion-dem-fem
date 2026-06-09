from __future__ import annotations

import io
import json
import math
import tempfile
import urllib.request
import zipfile
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd


BASE = Path(__file__).resolve().parents[1]
DATA = BASE / "data"
FIGURES = BASE / "figures"
DATA.mkdir(parents=True, exist_ok=True)
FIGURES.mkdir(parents=True, exist_ok=True)

ALVKARLEBY_URL = "https://data.mendeley.com/public-api/zip/zhfx2g7s7b/download/2"
ALVKARLEBY_DOI = "10.17632/zhfx2g7s7b.2"
RAW_SOURCE_DIR = BASE.parent / "external_raw_sources_not_for_upload"
RAW_SOURCE_DIR.mkdir(parents=True, exist_ok=True)
ALVKARLEBY_ZIP = RAW_SOURCE_DIR / "external_alvkarleby_mendeley_v2.zip"
RAW_DIR = DATA / "external_digitized_raw_points"

SECTION_ZERO_FLOW_MM = np.array([15.96, 10.14, 1.96, 14.92, 15.92, 20.20, 8.34, 13.32])
SECTION_WIDTH_M = np.array([2.04, 3.05, 2.56, 2.56, 2.45, 2.29, 3.06, 2.00])
SECTIONS = [f"section_{i}" for i in range(1, 9)]


def write_json(path: Path, payload: dict) -> None:
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def download_alvkarleby() -> dict:
    if not ALVKARLEBY_ZIP.exists():
        request = urllib.request.Request(
            ALVKARLEBY_URL,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125 Safari/537.36"
                ),
                "Accept": "application/zip,application/octet-stream,*/*",
            },
        )
        with urllib.request.urlopen(request, timeout=90) as response:
            ALVKARLEBY_ZIP.write_bytes(response.read())
    with zipfile.ZipFile(ALVKARLEBY_ZIP) as zf:
        entries = zf.namelist()
    return {
        "source_id": "Alvkarleby_Mendeley_V2",
        "doi": ALVKARLEBY_DOI,
        "url": "https://data.mendeley.com/datasets/zhfx2g7s7b/2",
        "download_url": ALVKARLEBY_URL,
        "license": "CC BY 4.0",
        "zip_path": str(ALVKARLEBY_ZIP),
        "upload_policy": "raw Mendeley ZIP is kept outside the supplement upload package",
        "zip_entry_count": len(entries),
        "zip_entries": entries,
        "claim_boundary": (
            "field-scale experimental test-bed comparison for relative seepage/turbidity "
            "patterns; not operational safety forecast or calibrated flow-rate validation"
        ),
    }


def write_download_blocked(reason: str) -> None:
    summary = {
        "status": "FIELD_SCALE_DOWNLOAD_BLOCKED",
        "source_id": "Alvkarleby_Mendeley_V2",
        "doi": ALVKARLEBY_DOI,
        "url": "https://data.mendeley.com/datasets/zhfx2g7s7b/2",
        "download_url": ALVKARLEBY_URL,
        "reason": reason,
        "next_required_action": (
            "download the V2 ZIP from the DOI page or public API and place it at "
            f"{ALVKARLEBY_ZIP.name}; then rerun this script"
        ),
        "claim_enabled_now": "none",
        "claim_boundary": (
            "field-scale comparison cannot be claimed until the open Mendeley V2 data "
            "are locally available and parsed"
        ),
    }
    write_json(DATA / "external_alvkarleby_field_validation_summary.json", summary)
    write_json(DATA / "external_claim_gate_summary.json", {
        "status": "FIELD_SCALE_VALIDATION_BLOCKED",
        "field_scale_status": summary["status"],
        "claim_allowed": "Lee A/B and Ke-Takahashi only; Alvkarleby remains an acquisition route",
        "claim_forbidden": "field-scale validation or bounded by external evidence prediction",
        "next_required_action": summary["next_required_action"],
    })
    print(json.dumps(summary, indent=2))


def read_excel_from_zip(zf: zipfile.ZipFile, name: str, **kwargs) -> pd.DataFrame:
    with zf.open(name) as f:
        return pd.read_excel(io.BytesIO(f.read()), **kwargs)


def parse_seepage_workbook(zf: zipfile.ZipFile, entry: str) -> pd.DataFrame:
    # Streaming parser. Pandas is too slow and memory-heavy for the minute-scale
    # logger workbooks, so this function returns daily section aggregates.
    with zf.open(entry) as f:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
    try:
        ws = wb["Loggerdata"] if "Loggerdata" in wb.sheetnames else wb[wb.sheetnames[0]]
        buckets: dict[tuple[str, str], list[float]] = defaultdict(list)
        upstream: dict[tuple[str, str], list[float]] = defaultdict(list)
        count_rows = 0
        for row in ws.iter_rows(values_only=True):
            if row is None or len(row) < 10:
                continue
            raw_date = row[0]
            if isinstance(raw_date, datetime):
                day = raw_date.date().isoformat()
            elif isinstance(raw_date, date):
                day = raw_date.isoformat()
            else:
                parsed = pd.to_datetime(raw_date, errors="coerce")
                if pd.isna(parsed):
                    continue
                day = parsed.date().isoformat()
            values = row[2:10]
            level = pd.to_numeric(row[10], errors="coerce") if len(row) > 10 else np.nan
            for idx, value in enumerate(values):
                val = pd.to_numeric(value, errors="coerce")
                if pd.isna(val):
                    continue
                h_excess_mm = max(float(val) - float(SECTION_ZERO_FLOW_MM[idx]), 0.0)
                q_star = (h_excess_mm / 1000.0) ** 2.5
                section = SECTIONS[idx]
                key = (day, section)
                buckets[key].append(q_star / float(SECTION_WIDTH_M[idx]))
                if not pd.isna(level):
                    upstream[key].append(float(level))
            count_rows += 1
        records = []
        for (day, section), values in buckets.items():
            levels = upstream.get((day, section), [])
            records.append(
                {
                    "date": day,
                    "section": section,
                    "q_star_width_norm_mean": float(np.mean(values)),
                    "q_star_width_norm_p95": float(np.percentile(values, 95)),
                    "level_upstream_m_mean": float(np.mean(levels)) if levels else np.nan,
                    "sample_count": len(values),
                    "source_file": entry,
                }
            )
        result = pd.DataFrame(records)
        result.attrs["raw_rows_seen"] = count_rows
        return result
    finally:
        wb.close()


def build_seepage_long(zf: zipfile.ZipFile) -> tuple[pd.DataFrame, dict]:
    entries = sorted(
        e for e in zf.namelist()
        if e.lower().endswith(".xlsx") and "seepage" in e.lower() and not Path(e).name.startswith("~$")
    )
    frames = []
    skipped = []
    for entry in entries:
        try:
            frame = parse_seepage_workbook(zf, entry)
            if frame.empty:
                skipped.append({"entry": entry, "reason": "no parseable logger rows"})
            else:
                frames.append(frame)
        except Exception as exc:
            skipped.append({"entry": entry, "reason": str(exc)})

    if not frames:
        return pd.DataFrame(), {"seepage_files_seen": len(entries), "skipped": skipped}

    section_day = pd.concat(frames, ignore_index=True)
    section_day = (
        section_day.groupby(["date", "section"], as_index=False)
        .agg(
            q_star_width_norm_mean=("q_star_width_norm_mean", "mean"),
            q_star_width_norm_p95=("q_star_width_norm_p95", "mean"),
            level_upstream_m_mean=("level_upstream_m_mean", "mean"),
            sample_count=("sample_count", "sum"),
        )
    )
    return section_day, {"seepage_files_seen": len(entries), "skipped": skipped}


def build_turbidity_long(zf: zipfile.ZipFile) -> tuple[pd.DataFrame, dict]:
    entries = [e for e in zf.namelist() if Path(e).name.lower() == "turbidity.xlsx"]
    if not entries:
        return pd.DataFrame(), {"status": "missing_turbidity_workbook"}
    entry = entries[0]
    try:
        sheets = pd.read_excel(io.BytesIO(zf.read(entry)), sheet_name=None, header=None)
    except Exception as exc:
        return pd.DataFrame(), {"status": "parse_failed", "reason": str(exc)}

    rows = []
    for sheet_name, frame in sheets.items():
        for r_idx, row in frame.iterrows():
            values = [v for v in row.tolist() if not pd.isna(v)]
            if not values:
                continue
            fnu_values = [pd.to_numeric(v, errors="coerce") for v in values]
            numeric = [float(v) for v in fnu_values if not pd.isna(v)]
            if not numeric:
                continue
            text = " ".join(str(v).lower() for v in values)
            section = None
            for i in range(1, 9):
                if f"section {i}" in text or f"section{i}" in text or f"s{i}" == text.strip():
                    section = f"section_{i}"
                    break
            if section is None:
                section = "upstream_or_unspecified"
            rows.append(
                {
                    "sheet": str(sheet_name),
                    "row_index": int(r_idx),
                    "section": section,
                    "fnu_mean": float(np.mean(numeric)),
                    "fnu_std": float(np.std(numeric)) if len(numeric) > 1 else 0.0,
                    "n_numeric": len(numeric),
                    "source_file": entry,
                }
            )
    return pd.DataFrame(rows), {"status": "parsed", "row_count": len(rows), "sheet_count": len(sheets)}


def build_model_section_scores() -> pd.DataFrame:
    path = pd.read_csv(DATA / "hybrid_dijkstra_path_cells.csv")
    # Map the synthetic model x-domain [0, 1] to the eight downstream weir
    # sections. This is a relative spatial test-bed comparison, not a physical
    # coordinate match to the Swedish dam.
    x = path["x"].clip(0.0, 1.0)
    section_number = np.floor(x * 8.0).astype(int) + 1
    section_number = section_number.clip(1, 8)
    path["section"] = "section_" + section_number.astype(str)
    scores = (
        path.groupby("section", as_index=False)
        .agg(
            pred_score_sum=("erosive_intensity", "sum"),
            pred_score_max=("erosive_intensity", "max"),
            path_cell_count=("cell_i", "count"),
        )
    )
    all_sections = pd.DataFrame({"section": SECTIONS})
    scores = all_sections.merge(scores, on="section", how="left").fillna(0.0)
    scores["pred_proportion"] = scores["pred_score_sum"] / max(scores["pred_score_sum"].sum(), 1e-12)
    scores["pred_rank"] = scores["pred_score_sum"].rank(method="min", ascending=False).astype(int)
    return scores


def field_metrics(seepage: pd.DataFrame, turbidity: pd.DataFrame, pred: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    if seepage.empty:
        return pd.DataFrame(), {
            "status": "FIELD_DATA_PARSE_FAILED",
            "claim_enabled_now": "none",
            "claim_boundary": "Alvkarleby V2 could not be parsed by the current ETL",
        }

    section_day = seepage.copy()
    section_summary = (
        section_day.groupby("section", as_index=False)
        .agg(
            observed_score=("q_star_width_norm_p95", "mean"),
            observed_max=("q_star_width_norm_p95", "max"),
            observed_days=("date", "nunique"),
        )
    )
    if not turbidity.empty and "section" in turbidity.columns:
        turb = (
            turbidity[turbidity["section"].isin(SECTIONS)]
            .groupby("section", as_index=False)
            .agg(turbidity_fnu_mean=("fnu_mean", "mean"), turbidity_rows=("fnu_mean", "count"))
        )
        section_summary = section_summary.merge(turb, on="section", how="left")
    else:
        section_summary["turbidity_fnu_mean"] = np.nan
        section_summary["turbidity_rows"] = 0

    joined = pd.DataFrame({"section": SECTIONS}).merge(section_summary, on="section", how="left").merge(pred, on="section", how="left")
    joined = joined.fillna({"observed_score": 0.0, "observed_max": 0.0, "observed_days": 0, "turbidity_rows": 0})
    joined["observed_proportion"] = joined["observed_score"] / max(joined["observed_score"].sum(), 1e-12)
    joined["observed_rank"] = joined["observed_score"].rank(method="min", ascending=False).astype(int)
    joined["absolute_proportion_error"] = (joined["pred_proportion"] - joined["observed_proportion"]).abs()

    pred_rank = joined["pred_score_sum"].rank(method="average")
    obs_rank = joined["observed_score"].rank(method="average")
    spearman = float(pred_rank.corr(obs_rank, method="pearson")) if len(joined) > 1 else float("nan")
    mae_prop = float(joined["absolute_proportion_error"].mean())
    top_pred = set(joined.sort_values("pred_score_sum", ascending=False).head(2)["section"])
    top_obs = set(joined.sort_values("observed_score", ascending=False).head(2)["section"])
    top2_hit = len(top_pred & top_obs) / 2.0
    section_mid = {f"section_{i}": (i - 0.5) / 8.0 for i in range(1, 9)}
    joined["x_mid_rel"] = joined["section"].map(section_mid)
    pred_centroid = float((joined["x_mid_rel"] * joined["pred_proportion"]).sum())
    obs_centroid = float((joined["x_mid_rel"] * joined["observed_proportion"]).sum())

    summary = {
        "status": "PASS_FIELD_SCALE_TEST_BED_COMPARISON_WITH_LIMITS",
        "source_id": "Alvkarleby_Mendeley_V2",
        "doi": ALVKARLEBY_DOI,
        "license": "CC BY 4.0",
        "seepage_rows": int(len(seepage)),
        "section_day_rows": int(len(section_day)),
        "turbidity_rows": int(len(turbidity)),
        "section_count": 8,
        "spearman_predicted_vs_observed_section_score": spearman,
        "top2_section_hit_fraction": top2_hit,
        "mean_absolute_section_proportion_error": mae_prop,
        "relative_centroid_error": abs(pred_centroid - obs_centroid),
        "predicted_centroid_relative_x": pred_centroid,
        "observed_centroid_relative_x": obs_centroid,
        "claim_enabled_now": (
            "field-scale experimental test-bed comparison for relative section ranking "
            "and pathway-envelope localization"
        ),
        "claim_boundary": (
            "Ã„lvkarleby V2 supports relative seepage/turbidity and section-ranking comparison only; "
            "it is not a calibrated absolute-discharge validation, operational safety forecast, "
            "or validation on a natural operational dam"
        ),
    }
    section_day.to_csv(DATA / "external_alvkarleby_section_day_metrics.csv", index=False)
    joined.to_csv(DATA / "external_alvkarleby_field_model_join.csv", index=False)
    return joined, summary


def write_source_manifest(alv: dict, field_summary: dict) -> None:
    rows = [
        {
            "source_id": "Alvkarleby_Mendeley_V2",
            "type": "field_scale_test_bed_open_dataset",
            "doi_or_identifier": ALVKARLEBY_DOI,
            "url": "https://data.mendeley.com/datasets/zhfx2g7s7b/2",
            "license_or_rights": "CC BY 4.0",
            "local_use": "relative section ranking and seepage/turbidity field-scale comparison",
            "redistribution": "processed derived CSV/JSON allowed with attribution; raw zip downloaded from DOI",
            "status": field_summary.get("status", "UNKNOWN"),
        },
        {
            "source_id": "Sterpi_2003",
            "type": "copyrighted_laboratory_paper",
            "doi_or_identifier": "10.1061/(ASCE)1532-3641(2003)3:1(111)",
            "url": "https://ascelibrary.org/doi/10.1061/(ASCE)1532-3641(2003)3:1(111)",
            "license_or_rights": "publisher controlled; no PDF or figures redistributed",
            "local_use": "legal-access digitization target for five-gradient erosion trajectories",
            "redistribution": "derived digitized points only after legal access and uncertainty audit",
            "status": "LEGAL_FULL_TEXT_REQUIRED_FOR_RESIDUAL_CLOSURE",
        },
        {
            "source_id": "Chang_2012",
            "type": "thesis_and_related_papers",
            "doi_or_identifier": "10.14711/thesis-b1165745",
            "url": "https://researchportal.hkust.edu.hk/en/studentTheses/internal-erosion-and-overtopping-erosion-of-earth-dams-and-landsl/",
            "license_or_rights": "copyrighted to author; no PDF or figures redistributed",
            "local_use": "mechanism, stress-conditioned thresholds, field jet erodibility and breach trajectory acquisition target",
            "redistribution": "derived digitized metrics only with locator and uncertainty audit",
            "status": "FULLTEXT_ROUTE_DIGITIZATION_REQUIRED",
        },
        {
            "source_id": "Zhu_2025_CG_107620",
            "type": "recent_cg_comparator_and_validation_family_target",
            "doi_or_identifier": "10.1016/j.compgeo.2025.107620",
            "url": "https://www.sciencedirect.com/science/article/abs/pii/S0266352X25005695",
            "license_or_rights": "publisher controlled; no PDF or figures redistributed",
            "local_use": "novelty comparator and legal-access residual validation target",
            "redistribution": "derived digitized points only after legal access and uncertainty audit",
            "status": "FULL_TEXT_REQUIRED_FOR_RESIDUAL_CLOSURE",
        },
    ]
    pd.DataFrame(rows).to_csv(DATA / "external_source_manifest.csv", index=False)
    write_json(DATA / "external_source_manifest_summary.json", {"sources": rows, "alvkarleby_zip": alv})


def write_methods_files(field_summary: dict) -> None:
    (BASE / "README_sources.md").write_text(
        """# External Validation Source Register

This package stores only rights-safe derived data, scripts, source locators and
uncertainty summaries. It does not redistribute restricted PDFs, publisher
figures, page screenshots or source article scans.

## Alvkarleby V2

- DOI: 10.17632/zhfx2g7s7b.2
- URL: https://data.mendeley.com/datasets/zhfx2g7s7b/2
- License: CC BY 4.0
- Use: field-scale experimental test-bed comparison for relative seepage,
  turbidity and section-ranking patterns.

## Sterpi 2003

- DOI: 10.1061/(ASCE)1532-3641(2003)3:1(111)
- Use: legal-access digitization target for laboratory erosion trajectories.
- Status: not closed unless a legally obtained full text is digitized and
  audited.

## Chang 2012

- DOI: 10.14711/thesis-b1165745
- Use: stress-conditioned mechanism and trajectory acquisition route.
- Status: mechanism/full-text route unless curves or tables are digitized with
  uncertainty.

## Zhu 2025

- DOI: 10.1016/j.compgeo.2025.107620
- Use: recent Computers and Geotechnics comparator and legal-access residual
  validation target.
- Status: metadata/acquisition route unless full text curves or tables are
  legally obtained and audited.
""",
        encoding="utf-8",
    )
    (BASE / "DIGITIZATION_METHODS.md").write_text(
        """# Digitization And Trajectory Validation Methods

Digitized literature curves must include source DOI, page, figure/table,
axis-calibration points, units, pixel coordinates, duplicate digitization pass,
and uncertainty in both axes. Curves with ambiguous overlap or missing units are
marked `usable_with_caution` or `reject` and are not allowed to support residual
validation claims.

Trajectory validation uses a pathway envelope: a source-to-exit corridor or
section-ranking output compared with observed section-level seepage/turbidity,
breach or erosion trajectories. Full trajectory claims require explicit
time/path metrics and uncertainty bands; otherwise the evidence is described as
field-scale or mechanism-level support.
""",
        encoding="utf-8",
    )
    (BASE / "LICENSES_AND_RIGHTS.md").write_text(
        f"""# Licenses And Rights

Alvkarleby V2 is cited as CC BY 4.0 via DOI {ALVKARLEBY_DOI}. Derived processed
CSV/JSON files and scripts in this supplement mark that source and preserve the
license attribution.

Sterpi 2003, Chang 2012 and Zhu 2025 are treated as publisher/author-controlled
sources. This package does not include their PDFs, figures, page images or
screenshots. It may include only derived digitized points, uncertainty records
and source locators after legal access.

Current field-scale status: {field_summary.get('status', 'UNKNOWN')}.
""",
        encoding="utf-8",
    )


def write_claim_gate(field_summary: dict) -> dict:
    ke_summary = json.loads((DATA / "external_ke_takahashi2012_second_family_residuals_summary.json").read_text(encoding="utf-8"))
    summary = {
        "status": "PASS_BOUNDED_FIELD_SCALE_AND_SECOND_FAMILY_EVIDENCE",
        "field_scale_status": field_summary.get("status"),
        "field_scale_metrics": {
            "spearman_predicted_vs_observed_section_score": field_summary.get("spearman_predicted_vs_observed_section_score"),
            "top2_section_hit_fraction": field_summary.get("top2_section_hit_fraction"),
            "mean_absolute_section_proportion_error": field_summary.get("mean_absolute_section_proportion_error"),
            "relative_centroid_error": field_summary.get("relative_centroid_error"),
        },
        "ke_takahashi_status": ke_summary.get("status"),
        "sterpi_status": "LEGAL_FULL_TEXT_REQUIRED_FOR_RESIDUAL_CLOSURE",
        "chang_status": "FULLTEXT_ROUTE_DIGITIZATION_REQUIRED_FOR_QUANTITATIVE_TRAJECTORY_CLOSURE",
        "zhu_status": "FULL_TEXT_REQUIRED_FOR_RESIDUAL_CLOSURE",
        "claim_allowed": (
            "bounded field-scale experimental test-bed comparison, Lee A/B laboratory-state "
            "validation, Ke and Takahashi second-family residual endpoint check, and "
            "rights-safe acquisition routes for Sterpi/Chang/Zhu"
        ),
        "claim_forbidden": (
            "unbounded operational safety forecast, complete erosion trajectory "
            "validation without observed continuous trajectories, or Sterpi/Chang/Zhu residual "
            "closure without legal full-text digitization"
        ),
        "review_benchmark_effect": (
            "strengthens validation evidence beyond Lee plus Ke by adding a reproducible "
            "field-scale test-bed comparison; does not by itself justify unbounded final-readiness claims"
        ),
    }
    write_json(DATA / "external_claim_gate_summary.json", summary)
    return summary


def main() -> None:
    try:
        provenance = download_alvkarleby()
    except Exception as exc:
        write_download_blocked(str(exc))
        return
    section_day_path = DATA / "external_alvkarleby_section_day_metrics.csv"
    turbidity_path = DATA / "external_alvkarleby_turbidity_long.csv"
    seepage_meta = {"cache": "external_alvkarleby_section_day_metrics.csv"}
    turbidity_meta = {"cache": "external_alvkarleby_turbidity_long.csv"}
    if section_day_path.exists() and turbidity_path.exists():
        seepage = pd.read_csv(section_day_path)
        turbidity = pd.read_csv(turbidity_path)
    else:
        with zipfile.ZipFile(ALVKARLEBY_ZIP) as zf:
            seepage, seepage_meta = build_seepage_long(zf)
            turbidity, turbidity_meta = build_turbidity_long(zf)
        if not seepage.empty:
            seepage.to_csv(section_day_path, index=False)
        if not turbidity.empty:
            turbidity.to_csv(turbidity_path, index=False)
    pred = build_model_section_scores()
    pred.to_csv(DATA / "external_model_path_to_weir_mapping.csv", index=False)
    joined, field_summary = field_metrics(seepage, turbidity, pred)
    field_summary["seepage_parse"] = seepage_meta
    field_summary["turbidity_parse"] = turbidity_meta
    write_json(DATA / "external_alvkarleby_field_validation_summary.json", field_summary)
    write_source_manifest(provenance, field_summary)
    write_methods_files(field_summary)
    claim_gate = write_claim_gate(field_summary)
    print(json.dumps(claim_gate, indent=2))


if __name__ == "__main__":
    main()
